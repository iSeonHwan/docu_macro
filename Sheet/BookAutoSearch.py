import pandas as pd
import requests
from bs4 import BeautifulSoup
import re
import time
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink

# 📁 입력 및 출력 파일 설정
INPUT_FILE = "book_titles_only.xlsx"
OUTPUT_FILE = "도서_자동완성_YES24_하이퍼링크버전.xlsx"

# 🧠 YES24 검색 함수
def search_yes24(title):
    """
    도서명(title)을 YES24에서 검색한 후,
    가격, 출판사, 발행연도, URL을 추출하여 반환합니다.
    """
    base_url = "https://www.yes24.com/Product/Search?domain=ALL&query="
    search_url = base_url + requests.utils.quote(title)
    headers = {"User-Agent": "Mozilla/5.0"}

    try:
        res = requests.get(search_url, headers=headers, timeout=10)
        soup = BeautifulSoup(res.text, "html.parser")

        # 🔍 첫 번째 도서 정보 찾기
        item = soup.select_one("ul#yesSchList li")  # 도서 리스트의 첫 항목

        if not item:
            return None

        # 💰 가격 추출
        price_tag = item.select_one(".yes_b")  # class="yes_b"에 가격 존재
        price = int(re.sub(r"[^\d]", "", price_tag.text)) if price_tag else None

        # 🔗 도서 상세 링크 추출
        link_tag = item.select_one("a.gd_name")
        book_url = "https://www.yes24.com" + link_tag["href"] if link_tag else ""

        # 📚 출판사와 발행일 추출
        detail_tag = item.select_one(".authPub.info_auth")  # 저자/출판사/출간일 섞인 태그
        pub_text = detail_tag.text.strip() if detail_tag else ""
        publisher, pub_year = "", ""

        # 예시: "손현주 저 | 특별한서재 | 2021년 10월 15일"
        parts = [p.strip() for p in pub_text.split("|")]
        if len(parts) >= 2:
            publisher = parts[1]
        if len(parts) >= 3:
            match = re.search(r"\d{4}", parts[2])
            pub_year = match.group() if match else ""

        return {
            "예상단가": price,
            "출판사": publisher,
            "발행 연도": pub_year,
            "출처(URL)": book_url,
            "비고": "성공"
        }

    except Exception as e:
        print(f"❌ 오류: {e}")
        return None

# 📥 엑셀 파일 불러오기
df = pd.read_excel(INPUT_FILE)
df.columns = df.columns.str.strip().str.replace(" ", "").str.replace("\n", "")  # 열 이름 정리

results = []

print(f"\n📚 총 {len(df)}권 자동 검색 시작...\n")

# 🔁 각 도서에 대해 YES24 검색
for idx, row in df.iterrows():
    title = str(row["신청도서명"]).strip()
    print(f"[{idx+1}] '{title}' 검색 중...")

    result = search_yes24(title)
    if result:
        row["예상단가"] = result["예상단가"]
        row["출판사"] = result["출판사"]
        row["발행 연도"] = result["발행 연도"]
        row["출처(URL)"] = result["출처(URL)"]
        row["비고"] = result["비고"]
        print(f"   ✅ 가격: {result['예상단가']}원 | 출판사: {result['출판사']} | 연도: {result['발행 연도']}")
    else:
        row["비고"] = "검색 실패"
        print("   ⚠️ 검색 실패!")

    results.append(row)
    time.sleep(1)  # 과부하 방지

# 💾 하이퍼링크 포함하여 저장
df_out = pd.DataFrame(results)

# 엑셀 쓰기 + 하이퍼링크 포함
wb = load_workbook(INPUT_FILE)
ws = wb.active
ws.title = "자동완성결과"

# 기존 내용 제거
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.value = None

# 새 내용 추가
for r_idx, row in enumerate(dataframe_to_rows(df_out, index=False, header=True), start=1):
    for c_idx, value in enumerate(row, start=1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        if r_idx > 1 and df_out.columns[c_idx - 1] == "출처(URL)" and isinstance(value, str) and value.startswith("http"):
            cell.hyperlink = value
            cell.font = Font(color="0000FF", underline="single")

wb.save(OUTPUT_FILE)
print(f"\n✅ 완료! '{OUTPUT_FILE}' 파일로 저장되었습니다.")
